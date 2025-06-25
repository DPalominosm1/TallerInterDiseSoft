import openpyxl
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
import os
import datetime

# Ajusta esta ruta si es necesario
EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'data', 'hospital.xlsx')


class CitasView(APIView):
    def obtener_citas(self):
        workbook = openpyxl.load_workbook(EXCEL_PATH)
        sheet = workbook['pacientes']  # Asegúrate de que esta sea la pestaña correcta
        citas = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cita = {
                "paciente_id": str(row[0]),
                "nombre": row[1],
                "apellido_paterno": row[2],
                "apellido_materno": row[3],
                "rut": row[4],
                "fecha": str(row[5]),
                "hora": str(row[6]),
                "estado": row[7],
                "doctor_id": str(row[8]) if len(row) > 8 else None
            }
            citas.append(cita)
        return citas, workbook, sheet

    def obtener_doctor_por_fecha_hora(self, fecha_str, hora_str):
        # Convierte fecha string a objeto datetime para saber día de la semana
        fecha_dt = datetime.datetime.strptime(fecha_str, "%Y-%m-%d")
        dias_semana = ['lunes', 'martes', 'miércoles', 'jueves', 'viernes', 'sabado', 'domingo']
        dia_semana = dias_semana[fecha_dt.weekday()]  # weekday: 0=lunes ... 6=domingo

        # Abre el archivo y la hoja de doctores
        workbook = openpyxl.load_workbook(EXCEL_PATH)
        sheet_doctores = workbook['doctores']

        for row in sheet_doctores.iter_rows(min_row=2, values_only=True):
            doctor_id = str(row[0])
            # Índices para días: lunes=5, martes=6, miércoles=7, jueves=8, viernes=9, sabado=10 (basado en el orden de columnas)
            dias_columnas = {
                'lunes': 5,
                'martes': 6,
                'miércoles': 7,
                'jueves': 8,
                'viernes': 9,
                'sabado': 10
            }
            if dia_semana not in dias_columnas:
                continue  # No asignamos doctor para domingos

            horario = row[dias_columnas[dia_semana]]  # ej: '08:00-9:30'
            if not horario:
                continue

            # Hora cita como datetime.time
            try:
                hora_cita = datetime.datetime.strptime(hora_str, "%H:%M").time()
            except Exception:
                return None

            # Parsear rango horario doctor
            try:
                hora_inicio_str, hora_fin_str = horario.split('-')
                hora_inicio = datetime.datetime.strptime(hora_inicio_str.strip(), "%H:%M").time()
                hora_fin = datetime.datetime.strptime(hora_fin_str.strip(), "%H:%M").time()
            except Exception:
                continue  # Si no se puede parsear el horario, saltar

            # Ver si la hora de la cita está dentro del rango horario del doctor
            if hora_inicio <= hora_cita < hora_fin:
                return doctor_id

        return None

    def get(self, request, *args, **kwargs):
        citas, _, _ = self.obtener_citas()
        return Response(citas)

    def post(self, request, *args, **kwargs):
        data = request.data

        # Obtener doctor_id automático
        doctor_id = self.obtener_doctor_por_fecha_hora(data.get("fecha"), data.get("hora"))

        citas, workbook, sheet = self.obtener_citas()

        new_row = [
            data.get("paciente_id"),
            data.get("nombre"),
            data.get("apellido_paterno"),
            data.get("apellido_materno"),
            data.get("rut"),
            data.get("fecha"),
            data.get("hora"),
            data.get("estado"),
            doctor_id
        ]

        sheet.append(new_row)
        workbook.save(EXCEL_PATH)

        # Incluir doctor_id en la respuesta
        response_data = data.copy()
        response_data["doctor_id"] = doctor_id

        return Response(response_data, status=status.HTTP_201_CREATED)


class CitaDetailView(APIView):
    def obtener_citas(self):
        workbook = openpyxl.load_workbook(EXCEL_PATH)
        sheet = workbook['pacientes']
        citas = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cita = {
                "paciente_id": str(row[0]),
                "nombre": row[1],
                "apellido_paterno": row[2],
                "apellido_materno": row[3],
                "rut": row[4],
                "fecha": str(row[5]),
                "hora": str(row[6]),
                "estado": row[7],
                "doctor_id": str(row[8]) if len(row) > 8 else None
            }
            citas.append(cita)
        return citas, workbook, sheet

    def obtener_doctor_por_fecha_hora(self, fecha_str, hora_str):
        # Misma función para consistencia
        fecha_dt = datetime.datetime.strptime(fecha_str, "%Y-%m-%d")
        dias_semana = ['lunes', 'martes', 'miércoles', 'jueves', 'viernes', 'sabado', 'domingo']
        dia_semana = dias_semana[fecha_dt.weekday()]

        workbook = openpyxl.load_workbook(EXCEL_PATH)
        sheet_doctores = workbook['doctores']

        for row in sheet_doctores.iter_rows(min_row=2, values_only=True):
            doctor_id = str(row[0])
            dias_columnas = {
                'lunes': 5,
                'martes': 6,
                'miércoles': 7,
                'jueves': 8,
                'viernes': 9,
                'sabado': 10
            }
            if dia_semana not in dias_columnas:
                continue

            horario = row[dias_columnas[dia_semana]]
            if not horario:
                continue

            try:
                hora_cita = datetime.datetime.strptime(hora_str, "%H:%M").time()
            except Exception:
                return None

            try:
                hora_inicio_str, hora_fin_str = horario.split('-')
                hora_inicio = datetime.datetime.strptime(hora_inicio_str.strip(), "%H:%M").time()
                hora_fin = datetime.datetime.strptime(hora_fin_str.strip(), "%H:%M").time()
            except Exception:
                continue

            if hora_inicio <= hora_cita < hora_fin:
                return doctor_id

        return None

    def put(self, request, paciente_id, *args, **kwargs):
        data = request.data
        _, workbook, sheet = self.obtener_citas()

        doctor_id = self.obtener_doctor_por_fecha_hora(data.get("fecha"), data.get("hora"))

        for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if str(row[0].value) == str(paciente_id):
                sheet.cell(row=i, column=2).value = data.get("nombre")
                sheet.cell(row=i, column=3).value = data.get("apellido_paterno")
                sheet.cell(row=i, column=4).value = data.get("apellido_materno")
                sheet.cell(row=i, column=5).value = data.get("rut")
                sheet.cell(row=i, column=6).value = data.get("fecha")
                sheet.cell(row=i, column=7).value = data.get("hora")
                sheet.cell(row=i, column=8).value = data.get("estado")
                sheet.cell(row=i, column=9).value = doctor_id
                workbook.save(EXCEL_PATH)
                return Response(data, status=status.HTTP_200_OK)
        return Response({"error": "Cita no encontrada"}, status=status.HTTP_404_NOT_FOUND)

    def delete(self, request, paciente_id, *args, **kwargs):
        _, workbook, sheet = self.obtener_citas()
        for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            if str(row[0].value) == str(paciente_id):
                sheet.delete_rows(i, 1)
                workbook.save(EXCEL_PATH)
                return Response(status=status.HTTP_204_NO_CONTENT)
        return Response({"error": "Cita no encontrada"}, status=status.HTTP_404_NOT_FOUND)


class ConsultasPorPacienteAPIView(APIView):
    def obtener_citas(self):
        workbook = openpyxl.load_workbook(EXCEL_PATH)
        sheet = workbook["pacientes"]
        citas = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cita = {
                "paciente_id": str(row[0]),
                "nombre": row[1],
                "apellido_paterno": row[2],
                "apellido_materno": row[3],
                "rut": row[4],
                "fecha": str(row[5]),
                "hora": str(row[6]),
                "estado": row[7],
                "doctor_id": str(row[8]) if len(row) > 8 else None
            }
            citas.append(cita)
        return citas

    def get(self, request, paciente_id, *args, **kwargs):
        citas = self.obtener_citas()
        consultas = [c for c in citas if str(c["paciente_id"]) == str(paciente_id)]
        return Response(consultas, status=status.HTTP_200_OK)
